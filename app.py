from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from PyPDF2 import PdfMerger, PdfReader
from docx2pdf import convert as docx_to_pdf
import io

def excel_to_pdf(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active

    pdf = canvas.Canvas('temp_excel.pdf', pagesize=letter)

    x_offset = 50
    y_offset = 750
    line_height = 20

    for row in ws.iter_rows():
        for cell in row:
            cell_value = str(cell.value)
            pdf.drawString(x_offset, y_offset, cell_value)
            y_offset -= line_height
        x_offset = 50
        y_offset -= 10

        if y_offset < 50:
            pdf.showPage()
            y_offset = 750

    pdf.save()

def convert_image_to_pdf(image_file):
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    c.drawImage(image_file, 0, 0, width=letter[0], height=letter[1])
    c.showPage()
    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer

# Convert Excel to PDF
excel_to_pdf('dsa.xlsx')

# Merge Excel PDF with images and existing PDFs
pdf_merger = PdfMerger()

pdf_merger.append('temp_excel.pdf')  # Append Excel PDF

image_filenames = ['img.jpg']  # Replace with your image filenames
for image_file in image_filenames:
    image_pdf_buffer = convert_image_to_pdf(image_file)
    pdf_merger.append(image_pdf_buffer)

existing_pdfs = ['output.pdf']  # Replace with your existing PDF filenames
for pdf_file in existing_pdfs:
    pdf_merger.append(pdf_file)

# Convert Word documents to PDFs and add to the PDF merger
word_files = ['doc.docx']  # Replace with your Word document filenames in docx format
for word_file in word_files:
    pdf_file = word_file.replace('.docx', '.pdf')
    docx_to_pdf(word_file, pdf_file)
    pdf_merger.append(pdf_file)

# Save the merged PDF
output_pdf = 'final.pdf'
with open(output_pdf, 'wb') as output:
    pdf_merger.write(output)
# save all the documents i.e. images, excel sheet,word docx , pdf in the directory same as that of project only.